import win32com.client as win32
import os
from Person import *
from Table import *
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font

#Close Excel Workbook to write to file
def close_excel_workbook(workbook_name):
    try:
        excel = win32.Dispatch("Excel.Application")
        for workbook in excel.Workbooks:
            if workbook.Name == workbook_name:
                workbook.Close(SaveChanges=True)
    except Exception as e:
        print("Error while closing the workbook:", str(e))


#ReOpen Excel Workbook after program
def open_specific_workbook(file_path):
    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = True
        workbook = excel.Workbooks.Open(file_path)
    except Exception as e:
        print("Error while opening the workbook:", str(e))


def run_excel_macro(workbook_path, macro_name):
    try:
        # Get a reference to the running Excel application
        excel_app = win32.GetActiveObject("Excel.Application")
    except Exception as e:
        print("Error: Failed to get the running Excel instance.")
        return

    # Open the workbook
    try:
        workbook = excel_app.Workbooks.Open(workbook_path)
    except Exception as e:
        print("Error: Failed to open the workbook.")
        return

    # Run the macro
    try:
        excel_app.Run("'" + workbook.Name + "'!" + macro_name)
    except Exception as e:
        print("Error: Failed to run the macro.")
        return

    # Save the workbook
    try:
        workbook.Save()
    except Exception as e:
        print("Error: Failed to save and close the workbook.")
        return


#Function to find index of person with highest tier in Guest List
def personMaxTier(arr):
    maxPerson = 0
    for i in range(1, len(arr)):
        if arr[i].getTier() < arr[maxPerson].getTier():
            maxPerson = i
    return maxPerson



if __name__ == "__main__":
    close_excel_workbook("Seating_Chart_App.xlsm")
    entreeExists = False
    showJobs = False

        
    #Open Workbook
    path = os.getcwd()  + "\Seating_Chart_App.xlsm"
    workbook = load_workbook(filename=path, keep_vba=True)
    data = workbook["Guest Input"]


    #Guest Retrieval
    Guests = []
    for row in data.iter_rows(min_row = 3, values_only=True):
        if row[0] == None:
            continue
        
        if(row[1] != None):
            Name = (row[0].strip() + " " + row[1].strip()).lower()
        else:
            Name = row[0].strip().lower()
        
        if(row[2] != None):
            Tier = row[2] + 1
        else:
            Tier = 4

        Job = row[3]
        if Job != None and data["K12"].value.lower() == 'y':
            showJobs = True


        if row[4] != None:
            themeStr = row[4].lower()
            Themes = [s.strip() for s in themeStr.split(',')]
        else:
            Themes = []

        if row[5] != None:
            friendStr = row[5].lower()
            Friends = [s.strip() for s in friendStr.split(',')]
        else:
            Friends = []

        if row[6] != None:
            enemyStr = row[6].lower()
            Enemies = [s.strip() for s in enemyStr.split(',')]
        else:
            Enemies = []

        if Friends != [] or Enemies != []:
            Tier -= 1

        if row[7] != None and data["K11"].value != None and data["K11"].value.lower() == 'y':
            Entree = row[7].upper()
            entreeExists = True
        else:
            Entree = None

        Guests.append(Person(Name, Tier, Job, Themes, Friends, Enemies, Entree))


    #Accounting for Empty Seats
    numberOfPeople = len(Guests)
    noTables = data["K2"].value
    noPplPerTable = data["K3"].value
    noEmptySeats = (noPplPerTable * noTables) - numberOfPeople

    if noEmptySeats > 0:
        for s in range(noEmptySeats):
            Guests.append(Person("Empty Seat " + str(s + 1)))
        numberOfPeople = len(Guests)
    

    #Establishing One-Way Tie
    for i in range(numberOfPeople):
        for x in range(numberOfPeople):
            if x != i:
                Guests[i].addPerson(Guests[x], noPplPerTable)

    
    #Establishing Two-Way Tie
    for i in range(numberOfPeople):
        for x in range(i + 1, numberOfPeople):
            initialScore = Guests[i].getScore(Guests[x].getName())
            Guests[i].addScore(Guests[x].getName(), Guests[x].getScore(Guests[i].getName()))
            Guests[x].addScore(Guests[i].getName(), initialScore)

    for person in Guests:
        print(person.getName() + " Relationship Scores -------------------------------")
        person.printRelationships()
        print('\n')


    #Table Organizational Logic
    tableMap = [Table(noPplPerTable, showJobs) for _ in range(noTables)]

    #Distribute Empty Seats if Necessary
    distributeSeats = data["K13"].value.lower() == 'y'
    if noEmptySeats > 0 and distributeSeats:
        tableCount = 0
        index = numberOfPeople - noEmptySeats
        for seat in range(noEmptySeats):
            tableMap[tableCount].addPerson(Guests[index], "Adding Initial Guest696969")
            tableCount += 1
            if tableCount == len(tableMap):
                tableCount = 0
            del Guests[index]


    #Distribute people amongst tables
    for table in tableMap:
        while not table.isFull():
            if table.isEmpty() or (table.containsOnlyEmpties() and distributeSeats):
                maxInd = personMaxTier(Guests)
                table.addPerson(Guests[maxInd], Guests[maxInd].getEntree(), Guests)
                del Guests[maxInd]
            else:
                personToAdd = table.personToAdd()
                for guest in range(len(Guests)):
                    if Guests[guest].getName() == personToAdd:
                        table.addPerson(Guests[guest], Guests[guest].getEntree())
                        del Guests[guest]
                        break
    

    
    #Outputting Tables with People sitting there
    output_sheet = workbook.create_sheet("Table Assignments")
    r = 1
    c = 1
    label = "Guest Names"
    if showJobs:
        label += ", Job/Position"
    
    if entreeExists:
        label += ", Preferred Entree"

    for i in range(len(tableMap)):
        output_sheet.cell(row = r, column = 1).value = label
        output_sheet.cell(row = r, column = 2).value = "Table Number"
        output_sheet.cell(row = r, column = 1).alignment = Alignment(horizontal='center')
        output_sheet.cell(row = r, column = 2).alignment = Alignment(horizontal='center')
        output_sheet.cell(row = r, column = 1).font = Font(bold=True)
        output_sheet.cell(row = r, column = 2).font = Font(bold=True)

        r += 1
        r = tableMap[i].printNames(r, c, output_sheet, i+1)
        r += 1
    
    output_sheet.column_dimensions['A'].width = 40  # Adjusts the width of column A to 40.
    output_sheet.column_dimensions['B'].width = 20  # Adjusts the width of column B to 20.
    workbook.save(filename=path)

    #Reopen Workbook
    open_specific_workbook(path)
    
    #Run Macro to Generate Chart Visuals of Seats
    run_excel_macro(path, "CreateDoughnutCharts")
 