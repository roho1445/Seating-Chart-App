'''
*****************************************************************************************************
Copyright (c) Rohith Venkatesh.                                                           
All Rights Reserved.      


Created by Rohith Venkatesh
UCLA EXTERNAL AFFAIRS INTERN SUMMER 2023

App created for UCLA External Affars and the Events team at the UCLA College

Contact:
rvenkatesh2025@ucla.edu
https://www.linkedin.com/in/rohithvenkatesh/
*****************************************************************************************************
'''




import os
import pyexcel as pe
import subprocess
import xlwings as xw
from Person import *
from Table import *
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from appscript import app, mactypes



def run_excel_macro(workbook_path, macro_name):
    # Connect to the running Excel application
    app = xw.apps.active

    # Replace "Seating_Chart_App.xlsm" with the actual name of your workbook, including the file extension
    workbook_name = "Seating_Chart_App.xlsm"

    # Reference the open workbook by name
    try:
        workbook = app.books[workbook_name]
        
        try:
            # Run the Excel macro using the workbook object
            workbook.macro(macro_name).run()
            print(f"The macro '{macro_name}' has been run in Excel.")
        except Exception as e:
            print(f"An error occurred: {e}")
    except KeyError:
        print(f"Workbook '{workbook_name}' not found.")

#Function to find index of person with highest tier in Guest List
def personMaxTier(arr):
    maxPerson = 0
    for i in range(1, len(arr)):
        if arr[i].getTier() < arr[maxPerson].getTier():
            maxPerson = i
    return maxPerson



if __name__ == "__main__":
   
    # Replace "Workbook Name" with the name of the workbook you want to close
    workbook_name = "Seating_Chart_App.xlsm"

    # AppleScript command to close the workbook in Excel
    applescript_command = f'tell application "Microsoft Excel" to close workbook "{workbook_name}" saving no'

    # Use the subprocess module to run the AppleScript command
    try:
        subprocess.run(["osascript", "-e", applescript_command], check=True)
        print(f"The workbook '{workbook_name}' has been closed.")
    except subprocess.CalledProcessError as e:
        print(f"An error occurred: {e}")


    #Open Workbook
    script_directory = os.path.dirname(os.path.abspath(__file__))
    path = script_directory + "/Seating_Chart_App.xlsm"
    workbook = load_workbook(filename=path, keep_vba=True)
    data = workbook["Guest Input"]

    entreeExists = False
    showJobs = False

        
    

    #Guest Retrieval
    Guests = []
    for row in data.iter_rows(min_row = 3, values_only=True):
        if row[0] == None:
            continue
        
        if(row[1] != None):
            Name = (row[0].strip() + " " + row[1].strip()).lower()
        else:
            Name = row[0].strip().lower()
        
        if(row[2] != None):
            Tier = row[2] + 1
        else:
            Tier = 4

        Job = row[3]
        if Job != None and data["K12"].value != None and data["K12"].value.lower() == 'y':
            showJobs = True


        if row[4] != None:
            themeStr = row[4].lower()
            Themes = [s.strip() for s in themeStr.split(',')]
        else:
            Themes = []

        if row[5] != None:
            friendStr = row[5].lower()
            Friends = [s.strip() for s in friendStr.split(',')]
        else:
            Friends = []

        if row[6] != None:
            enemyStr = row[6].lower()
            Enemies = [s.strip() for s in enemyStr.split(',')]
        else:
            Enemies = []

        if Friends != [] or Enemies != []:
            Tier -= 1

        if row[7] != None and data["K11"].value != None and data["K11"].value.lower() == 'y':
            Entree = row[7].upper()
            entreeExists = True
        else:
            Entree = None

        Guests.append(Person(Name, Tier, Job, Themes, Friends, Enemies, Entree))


    #Accounting for Empty Seats
    numberOfPeople = len(Guests)
    noTables = data["K2"].value
    noPplPerTable = data["K3"].value
    noEmptySeats = (noPplPerTable * noTables) - numberOfPeople

    if noEmptySeats > 0:
        for s in range(noEmptySeats):
            Guests.append(Person("Empty Seat " + str(s + 1)))
        numberOfPeople = len(Guests)
    

    #Establishing One-Way Tie
    for i in range(numberOfPeople):
        for x in range(numberOfPeople):
            if x != i:
                Guests[i].addPerson(Guests[x], noPplPerTable)

    
    #Establishing Two-Way Tie
    for i in range(numberOfPeople):
        for x in range(i + 1, numberOfPeople):
            initialScore = Guests[i].getScore(Guests[x].getName())
            Guests[i].addScore(Guests[x].getName(), Guests[x].getScore(Guests[i].getName()))
            Guests[x].addScore(Guests[i].getName(), initialScore)


    #Table Organizational Logic
    tableMap = [Table(noPplPerTable, showJobs) for _ in range(noTables)]

    #Distribute Empty Seats if Necessary
    distributeSeats = False
    if data["K13"].value != None:
        distributeSeats = data["K13"].value.lower() == 'y'
        if noEmptySeats > 0 and distributeSeats:
            tableCount = 0
            index = numberOfPeople - noEmptySeats
            for seat in range(noEmptySeats):
                tableMap[tableCount].addPerson(Guests[index], "Adding Initial Guest696969")
                tableCount += 1
                if tableCount == len(tableMap):
                    tableCount = 0
                del Guests[index]


    #Distribute people amongst tables
    for table in tableMap:
        while not table.isFull():
            if table.isEmpty() or (table.containsOnlyEmpties() and distributeSeats):
                maxInd = personMaxTier(Guests)
                table.addPerson(Guests[maxInd], Guests[maxInd].getEntree(), Guests)
                del Guests[maxInd]
            else:
                personToAdd = table.personToAdd()
                for guest in range(len(Guests)):
                    if Guests[guest].getName() == personToAdd:
                        table.addPerson(Guests[guest], Guests[guest].getEntree())
                        del Guests[guest]
                        break
    

    
    #Outputting Tables with People sitting there
    output_sheet = workbook.create_sheet("Table Assignments")
    r = 1
    c = 1
    label = "Guest Names"
    if showJobs:
        label += ", Job/Position"
    
    if entreeExists:
        label += ", Preferred Entree"

    for i in range(len(tableMap)):
        output_sheet.cell(row = r, column = 1).value = label
        output_sheet.cell(row = r, column = 2).value = "Table Number"
        output_sheet.cell(row = r, column = 1).alignment = Alignment(horizontal='center')
        output_sheet.cell(row = r, column = 2).alignment = Alignment(horizontal='center')
        output_sheet.cell(row = r, column = 1).font = Font(bold=True)
        output_sheet.cell(row = r, column = 2).font = Font(bold=True)

        r += 1
        r = tableMap[i].printNames(r, c, output_sheet, i+1)
        r += 1
    
    output_sheet.column_dimensions['A'].width = 40  # Adjusts the width of column A to 40.
    output_sheet.column_dimensions['B'].width = 20  # Adjusts the width of column B to 20.
    workbook.save(filename=path)

    # Build the command to open the workbook with Excel
    command = ["open", "-a", "Microsoft Excel", path]

    try:
        # Execute the command
        subprocess.run(command, check=True)
        print(f"The workbook '{path}' has been opened with Excel.")
    except subprocess.CalledProcessError as e:
        print(f"An error occurred: {e}")


    run_excel_macro(path, "CreateDoughnutCharts")
 