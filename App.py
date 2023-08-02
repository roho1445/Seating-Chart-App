from Person import *
from Table import *
from openpyxl import load_workbook
import win32com.client as win32

#Close Excel Workbook to write to file
def close_excel_workbook(file_path):
    try:
        excel = win32.Dispatch("Excel.Application")
        #workbook = excel.Workbooks(file_path)
        #workbook.Close(SaveChanges=False)
        excel.Quit()
    except Exception as e:
        print("Error while closing the workbook:", str(e))


#ReOpen Excel Workbook after program
def open_specific_workbook(file_path):
    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = True

        workbook = excel.Workbooks.Open(file_path)
        # Do operations with the workbook here if needed

    except Exception as e:
        print("Error while opening the workbook:", str(e))

#Function to find index of person with highest tier in Guest List
def personMaxTier(arr):
    maxPerson = 0
    for i in range(1, len(arr)):
        if arr[i].getTier() < arr[maxPerson].getTier():
            maxPerson = i
    return maxPerson



if __name__ == "__main__":

    close_excel_workbook(r"C:\Users\rohit\CS_Code\UCLA_EA_Internship\VSCode_Files\Seating_Chart_App.xlsm")
    #Open Workbook
    file = "Seating_Chart_App.xlsm"
    workbook = load_workbook(filename=file, keep_vba=True)
    data = workbook["Guest Input"]
    
    #Guest Retrieval
    Guests = []
    for row in data.iter_rows(min_row = 2, values_only=True):
        if row[0] == None:
            continue
        
        if(row[1] != None):
            Name = (row[0].strip() + " " + row[1].strip()).lower()
        else:
            Name = row[0].strip().lower()
        Tier = row[2]
        Job = row[3]
        

        if row[4] != None:
            themeStr = row[4].lower().replace(" ", "")
            Themes = themeStr.split(',')
        else:
            Themes = []

        if row[5] != None:
            friendStr = row[5].lower()
            Friends = [s.strip() for s in friendStr.split(',')]
        else:
            Friends = []

        if row[6] != None:
            enemyStr = row[6].lower()
            Enemies = [s.strip() for s in enemyStr.split(',')]
        else:
            Enemies = []

        Guests.append(Person(Name, Tier, Job, Themes, Friends, Enemies))
    

    #Accounting for Empty Seats
    numberOfPeople = len(Guests)
    noTables = data["K2"].value
    noPplPerTable = data["K3"].value

    if numberOfPeople < noPplPerTable * noTables:
        spares = (noPplPerTable * noTables) - numberOfPeople
        for s in range(spares):
            Guests.append(Person("Empty Seat " + str(s + 1)))
        numberOfPeople = len(Guests)
    

    #Establishing One-Way Tie
    for i in range(numberOfPeople):
        for x in range(numberOfPeople):
            if x != i:
                Guests[i].addPerson(Guests[x])
    
    #Establishing Two-Way Tie
    for i in range(numberOfPeople):
        for x in range(i + 1, numberOfPeople):
            initialScore = Guests[i].getScore(Guests[x].getName())

    
    #Table Organizational Logic
    tableMap = [Table(noPplPerTable) for _ in range(noTables)]

    for table in tableMap:
        while not table.isFull():
            if table.isEmpty():
                maxInd = personMaxTier(Guests)
                table.addPerson(Guests[maxInd], Guests)
                del Guests[maxInd]
            else:
                personToAdd = table.personToAdd()
                for guest in range(len(Guests)):
                    if Guests[guest].getName() == personToAdd:
                        table.addPerson(Guests[guest])
                        del Guests[guest]
                        break



    #Outputting Tables with People sitting there

    output_sheet = workbook.create_sheet("Table Assignments")
    r = 1
    c = 1
    for i in range(len(tableMap)):
        '''
        print("Table " + str(i+1) + " ----------")
        tableMap[i].printNames()
        print("\n")
        '''
        
        output_sheet.cell(row = r, column = c).value = "Table " + str(i+1) + " ----------"
        r += 1
        r = tableMap[i].printNames(r, c, output_sheet)
        r += 1


    workbook.save(filename=file)
    open_specific_workbook(r"C:\Users\rohit\CS_Code\UCLA_EA_Internship\VSCode_Files\Seating_Chart_App.xlsm")


    '''
    output_bk = Workbook()
    output_sheet = output_bk.active
    #output_sheet = workbook.create_sheet("Table Assignments")
    r = 1
    c = 1
    for i in range(len(tableMap)):
    
        print("Table " + str(i+1) + " ----------")
        tableMap[i].printNames()
        print("\n")
        
        
        output_sheet.cell(row = r, column = c).value = "Table " + str(i+1) + " ----------"
        r += 1
        r = tableMap[i].printNames(r, c, output_sheet)
        r += 1


    output_bk.save(filename="Table_Assignments.csv")
    '''

